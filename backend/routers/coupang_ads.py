import io

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from datetime import date

import store

router = APIRouter(prefix="/coupang-ads", tags=["coupang-ads"])

# 보고서2 (맞춤형 광고 보고서) 실제 컬럼명 매핑
_REQUIRED_COLS = [
    "광고 집행 옵션 ID",
    "광고집행 상품명",
    "캠페인 이름",
    "광고그룹",
    "노출수",
    "클릭수",
    "광고비(원)",
    "총 주문수 (14일)",
    "총 전환 매출액 (14일)(원)",
    "총 판매 수량 (14일)",
]


@router.post("/upload")
async def upload_wing_report(
    file: UploadFile = File(...),
    report_from: str = Form(...),
    report_to: str = Form(...),
):
    try:
        dt_from = date.fromisoformat(report_from)
        dt_to   = date.fromisoformat(report_to)
    except ValueError:
        raise HTTPException(400, "날짜 형식 오류 — YYYY-MM-DD")

    if dt_from > dt_to:
        raise HTTPException(400, "시작일이 종료일보다 늦음")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"파일 읽기 실패: {e}")

    headers = [cell.value for cell in ws[1]]
    col_idx = {name: i for i, name in enumerate(headers) if name in _REQUIRED_COLS}
    missing = [c for c in _REQUIRED_COLS if c not in col_idx]
    if missing:
        raise HTTPException(400, f"필수 컬럼 없음: {missing}")

    # vendorItemId 기준으로 지면별 행 합산
    aggregated: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vid_raw = row[col_idx["광고 집행 옵션 ID"]]
        if vid_raw is None:
            continue
        vid = str(int(vid_raw)) if isinstance(vid_raw, float) else str(vid_raw)

        if vid not in aggregated:
            aggregated[vid] = {
                "vendor_item_id": vid,
                "product_name":   row[col_idx["광고집행 상품명"]],
                "campaign_name":  row[col_idx["캠페인 이름"]],
                "adgroup_name":   row[col_idx["광고그룹"]],
                "impressions":    0,
                "clicks":         0,
                "ad_cost":        0,
                "orders_14d":     0,
                "conversion_revenue_14d": 0,
                "sales_qty_14d":  0,
            }

        agg = aggregated[vid]
        agg["impressions"] += int(row[col_idx["노출수"]] or 0)
        agg["clicks"]      += int(row[col_idx["클릭수"]] or 0)
        agg["ad_cost"]     += int(row[col_idx["광고비(원)"]] or 0)
        agg["orders_14d"]  += int(row[col_idx["총 주문수 (14일)"]] or 0)
        agg["conversion_revenue_14d"] += int(row[col_idx["총 전환 매출액 (14일)(원)"]] or 0)
        agg["sales_qty_14d"] += int(row[col_idx["총 판매 수량 (14일)"]] or 0)

    if not aggregated:
        raise HTTPException(400, "파싱된 데이터 없음")

    # vendorItemId → product_id 매핑 (최근 수집 기준)
    vid_map = store.get_vendor_item_id_map()

    records = []
    for vid, agg in aggregated.items():
        roas_14d = (
            round(agg["conversion_revenue_14d"] / agg["ad_cost"] * 100, 2)
            if agg["ad_cost"] > 0 else None
        )
        records.append({
            "report_from":            str(dt_from),
            "report_to":              str(dt_to),
            "vendor_item_id":         vid,
            "product_id":             vid_map.get(vid),
            "product_name":           agg["product_name"],
            "campaign_name":          agg["campaign_name"],
            "adgroup_name":           agg["adgroup_name"],
            "impressions":            agg["impressions"],
            "clicks":                 agg["clicks"],
            "ad_cost":                agg["ad_cost"],
            "orders_14d":             agg["orders_14d"],
            "conversion_revenue_14d": agg["conversion_revenue_14d"],
            "roas_14d":               roas_14d,
            "sales_qty_14d":          agg["sales_qty_14d"],
        })

    # 동일 기간 기존 데이터 삭제 후 재삽입
    store.delete_coupang_ad_report(str(dt_from), str(dt_to))
    store.save_coupang_ad_report(records)

    matched   = sum(1 for r in records if r["product_id"])
    unmatched = len(records) - matched

    return {
        "saved":       len(records),
        "matched":     matched,
        "unmatched":   unmatched,
        "report_from": str(dt_from),
        "report_to":   str(dt_to),
    }


@router.get("/summary")
def get_ad_summary():
    return store.get_latest_coupang_ad_summary()
